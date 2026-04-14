#!/usr/bin/env python3
"""
Solgar MINSAN scraper — farmaciaigea.com
VPS Hetzner Linux (Ubuntu/Debian)

Setup una tantum:
 pip install playwright openpyxl
 playwright install chromium
 playwright install-deps chromium

Uso:
 python3 solgar_scraper.py
 python3 solgar_scraper.py --output /percorso/custom.xlsx
"""

import asyncio
import re
import time
import argparse
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

from playwright.async_api import async_playwright, Page, TimeoutError as PWTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("solgar")

BASE_URL = "https://farmaciaigea.com"
# URL brand page con paginazione — mostra 36 prodotti/pagina
BRAND_URL = BASE_URL + "/marca/solgar?p={page}"

DELAY_PAGE = 2.5 # secondi tra pagine elenco
DELAY_PRODUCT = 1.5 # secondi tra schede prodotto


# ── STRUTTURA DATI ─────────────────────────────────────────────────────────────
@dataclass
class Prodotto:
 nome: str = ""
 minsan: str = ""
 ean: str = ""
 categoria: str = ""
 formato: str = ""
 composizione: str = ""
 url: str = ""
 prezzo: str = ""


# ── SCRAPING ELENCO ────────────────────────────────────────────────────────────
async def get_product_urls(page: Page) -> list[str]:
 """Raccoglie tutti gli URL prodotto dalla pagina brand, tutte le pagine."""
 urls = []
 page_num = 1

 while True:
 url = BRAND_URL.format(page=page_num)
 log.info(f"Elenco pagina {page_num}: {url}")

 try:
 await page.goto(url, wait_until="domcontentloaded", timeout=30_000)
 await page.wait_for_timeout(1500)
 except PWTimeout:
 log.warning(f"Timeout pagina {page_num}, salto")
 break

 # Raccoglie link prodotto
 links = await page.eval_on_selector_all(
 "a.product-thumbnail",
 "els => els.map(e => e.href)"
 )

 if not links:
 # Prova selettore alternativo
 links = await page.eval_on_selector_all(
 ".product_list .thumbnail a, h3.product-title a, .product-name a",
 "els => els.map(e => e.href)"
 )

 if not links:
 log.info(f"Nessun prodotto trovato a pagina {page_num}, fine paginazione")
 break

 # Filtra duplicati
 new_links = [l for l in links if l not in urls]
 urls.extend(new_links)
 log.info(f" → {len(new_links)} prodotti (totale: {len(urls)})")

 # Controlla se esiste pagina successiva
 next_btn = await page.query_selector("a[rel='next'], .pagination .next:not(.disabled)")
 if not next_btn:
 log.info("Nessuna pagina successiva, fine")
 break

 page_num += 1
 await asyncio.sleep(DELAY_PAGE)

 return urls


# ── SCRAPING SCHEDA PRODOTTO ────────────────────────────────────────────────────
async def scrape_product(page: Page, url: str) -> Optional[Prodotto]:
 try:
 await page.goto(url, wait_until="domcontentloaded", timeout=25_000)
 await page.wait_for_timeout(800)
 except PWTimeout:
 log.warning(f"Timeout: {url}")
 return None

 p = Prodotto(url=url)

 # Nome prodotto
 for sel in ["h1.product-detail-name", "h1[itemprop='name']", "h1"]:
 el = await page.query_selector(sel)
 if el:
 p.nome = (await el.inner_text()).strip()
 break

 # MINSAN / EAN — cerca in tutto il testo della pagina
 body_text = await page.inner_text("body")

 m = re.search(r"[Mm]in[Ss]an[:\s]*([0-9]{5,12})", body_text)
 if m:
 p.minsan = m.group(1).strip()

 e = re.search(r"EAN[:\s]*([0-9]{8,14})", body_text)
 if e:
 p.ean = e.group(1).strip()

 # Prezzo
 for sel in ["span.current-price", "[itemprop='price']", ".price"]:
 el = await page.query_selector(sel)
 if el:
 p.prezzo = (await el.inner_text()).strip().replace("\n", " ")
 break

 # Categoria — breadcrumb
 crumbs = await page.eval_on_selector_all(
 ".breadcrumb li, ol.breadcrumb span[itemprop='name']",
 "els => els.map(e => e.innerText.trim())"
 )
 if crumbs:
 # Salta home e nome prodotto, prendi la categoria intermedia
 cats = [c for c in crumbs if c.lower() not in ("", "home", p.nome.lower())]
 p.categoria = " > ".join(cats[-2:]) if cats else ""

 # Formato (es. "30 capsule", "100 tavolette") — cerca in description o features
 for sel in ["table.data-sheet td", ".product-features li", "#product-details"]:
 els = await page.query_selector_all(sel)
 for el in els:
 txt = (await el.inner_text()).strip()
 if re.search(r"\d+\s*(capsule|tavolette|perle|ml|g\b|softgel|bustine)", txt, re.IGNORECASE):
 p.formato = txt[:80]
 break
 if p.formato:
 break

 # Composizione / Ingredienti
 # Strategia 1: cerca sezione ingredienti/composizione esplicita
 comp_text = ""
 for sel in [
 "#ingredienti", "#composizione", "#composition",
 "div[id*='ingredi']", "div[id*='composiz']",
 ".product-description", "#product-description",
 "[itemprop='description']",
 ]:
 el = await page.query_selector(sel)
 if el:
 comp_text = (await el.inner_text()).strip()
 if comp_text:
 break

 # Strategia 2: cerca parole chiave nel testo completo
 if not comp_text:
 # Cerca paragrafi che contengono "ingredienti" o "composizione"
 paras = await page.eval_on_selector_all(
 "p, li",
 "els => els.map(e => e.innerText.trim())"
 )
 for para in paras:
 if re.search(r"ingredienti|composizione|contiene|antiagglomerante|agente di carica", para, re.IGNORECASE):
 comp_text = para
 break

 # Pulisce il testo composizione
 if comp_text:
 # Rimuove righe di disclaimer/avvertenze, taglia a 500 char
 lines = [l.strip() for l in comp_text.splitlines() if l.strip()]
 lines = [l for l in lines if not re.search(
 r"avverter|disclaimer|consult|bambini al di sotto|non superare|dose giornaliera",
 l, re.IGNORECASE
 )]
 p.composizione = " ".join(lines)[:500]

 log.info(f" ✓ {p.nome[:50]:<50} MINSAN={p.minsan or '—'}")
 return p


# ── EXPORT XLSX ────────────────────────────────────────────────────────────────
def export_xlsx(prodotti: list[Prodotto], output_path: str):
 wb = Workbook()
 ws = wb.active
 ws.title = "Solgar MINSAN"

 VERDE = "1A5C2A"
 BIANCO = "FFFFFF"
 GRIGIO = "F5F5F5"
 GIALLO = "FFF9C4"

 thin = Side(style="thin", color="CCCCCC")
 brd = Border(left=thin, right=thin, top=thin, bottom=thin)

 def hfill(c): return PatternFill("solid", fgColor=c)

 # Intestazione
 ws.merge_cells("A1:H1")
 c = ws["A1"]
 c.value = f"SOLGAR – Prodotti con MINSAN e Composizione ({len(prodotti)} referenze)"
 c.font = Font(name="Arial", bold=True, size=13, color=BIANCO)
 c.fill = hfill(VERDE)
 c.alignment = Alignment(horizontal="center", vertical="center")
 ws.row_dimensions[1].height = 26

 # Nota fonte
 ws.merge_cells("A2:H2")
 c = ws["A2"]
 c.value = "Fonte: farmaciaigea.com (scraping live) | Data raccolta: " + time.strftime("%d/%m/%Y")
 c.font = Font(name="Arial", italic=True, size=8, color="5D4037")
 c.fill = hfill(GIALLO)
 c.alignment = Alignment(horizontal="left", vertical="center")
 ws.row_dimensions[2].height = 16

 ws.row_dimensions[3].height = 5 # gap

 # Colonne
 cols = ["N°", "NOME PRODOTTO", "MINSAN", "EAN", "CATEGORIA", "FORMATO", "COMPOSIZIONE", "PREZZO"]
 widths = [5, 30, 13, 16, 28, 20, 70, 10]

 for ci, (h, w) in enumerate(zip(cols, widths), 1):
 c = ws.cell(row=4, column=ci, value=h)
 c.font = Font(name="Arial", bold=True, size=10, color=BIANCO)
 c.fill = hfill(VERDE)
 c.alignment = Alignment(horizontal="center", vertical="center")
 c.border = brd
 ws.column_dimensions[get_column_letter(ci)].width = w

 ws.row_dimensions[4].height = 18

 # Righe dati
 for i, p in enumerate(prodotti, 1):
 row = i + 4
 fill = hfill(BIANCO) if i % 2 == 0 else hfill(GRIGIO)
 vals = [i, p.nome, p.minsan, p.ean, p.categoria, p.formato, p.composizione, p.prezzo]
 for ci, val in enumerate(vals, 1):
 c = ws.cell(row=row, column=ci, value=val)
 c.font = Font(
 name="Arial", size=9,
 bold=(ci == 3 and bool(p.minsan)),
 color="1A5C2A" if ci == 3 and p.minsan else "000000"
 )
 c.fill = fill
 c.alignment = Alignment(
 vertical="top",
 wrap_text=(ci in (7,)),
 horizontal="center" if ci in (1, 3, 4) else "left"
 )
 c.border = brd
 ws.row_dimensions[row].height = 40

 ws.freeze_panes = "A5"
 ws.auto_filter.ref = f"A4:H{len(prodotti) + 4}"

 wb.save(output_path)
 log.info(f"Salvato: {output_path}")


# ── MAIN ───────────────────────────────────────────────────────────────────────
async def main(output_path: str):
 async with async_playwright() as pw:
 browser = await pw.chromium.launch(
 headless=True,
 args=[
 "--no-sandbox", # necessario su VPS senza display
 "--disable-setuid-sandbox",
 "--disable-dev-shm-usage", # evita crash per /dev/shm limitato
 "--disable-gpu",
 ]
 )
 context = await browser.new_context(
 user_agent=(
 "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
 "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
 ),
 locale="it-IT",
 viewport={"width": 1280, "height": 900},
 )
 page = await context.new_page()

 # 1. Raccoglie tutti gli URL prodotto
 log.info("=== FASE 1: raccolta URL prodotti ===")
 urls = await get_product_urls(page)
 log.info(f"Totale URL raccolti: {len(urls)}")

 if not urls:
 log.error("Nessun URL trovato. Controlla il selettore o l'URL base.")
 await browser.close()
 return

 # 2. Scraping schede prodotto
 log.info("=== FASE 2: scraping schede ===")
 prodotti = []
 for i, url in enumerate(urls, 1):
 log.info(f"[{i}/{len(urls)}] {url}")
 p = await scrape_product(page, url)
 if p:
 prodotti.append(p)
 await asyncio.sleep(DELAY_PRODUCT)

 # Checkpoint ogni 50 prodotti
 if i % 50 == 0:
 tmp = output_path.replace(".xlsx", f"_checkpoint_{i}.xlsx")
 export_xlsx(prodotti, tmp)
 log.info(f"Checkpoint salvato: {tmp}")

 await browser.close()

 # 3. Export finale
 log.info("=== FASE 3: export xlsx ===")
 export_xlsx(prodotti, output_path)

 # Riepilogo
 with_minsan = sum(1 for p in prodotti if p.minsan)
 log.info(f"=== COMPLETATO ===")
 log.info(f"Prodotti totali: {len(prodotti)}")
 log.info(f"Con MINSAN: {with_minsan}")
 log.info(f"Senza MINSAN: {len(prodotti) - with_minsan}")
 log.info(f"File: {output_path}")


if __name__ == "__main__":
 parser = argparse.ArgumentParser(description="Solgar MINSAN scraper")
 parser.add_argument("--output", default="Solgar_MINSAN_live.xlsx",
 help="Path file xlsx output (default: Solgar_MINSAN_live.xlsx)")
 args = parser.parse_args()
 asyncio.run(main(args.output))
